from openpyxl import Workbook

n = 21
res = []
for pct in range(0, 105, 5):
    pcsum = 0
    for i in range(200):
        f = open(f"./BPPercentage/newaquarium_10_{pct}_result/test{i}_res.txt", 'r')
        lines = f.readlines()
        f.close()
        pcsum += int(lines[0])
    res.append(pcsum/100)
write_wb = Workbook()
write_ws = write_wb.create_sheet('Result')
for i in range(n):
    write_ws.cell(2, i+5, res[i])
write_wb.save("GraphPerct.xlsx")
