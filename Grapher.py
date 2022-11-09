from openpyxl import Workbook

n = 400
res = [[] for _ in range(n)]
for i in range(100):
    f = open("./BPResult/aquarium_10_result/test"+str(i)+"_res.txt", 'r')
    lines = f.readlines()
    f.close()
    if len(lines)==0:
        continue
    for j in range(n):
        if j>=len(lines):
            if res[i][-1]<=10:
                res[i].append(0)
            else:
                res[i].append(res[i][-1])
        else:
            res[i].append(int(lines[j].split(" ")[0]))
write_wb = Workbook()
write_ws = write_wb.create_sheet('Result')
for i in range(100):
    for j in range(n):
        write_ws.cell(j+2, i+5, res[i][j])
write_wb.save("Graph.xlsx")
